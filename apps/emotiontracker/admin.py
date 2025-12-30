"""情绪追踪模块的 admin 配置"""
import io
from datetime import datetime, timedelta
from django.contrib import admin
from django.http import HttpResponse
from django.contrib import messages
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from apps.users.models import User
from apps.users.admin_mixins import UUIDUserAdminMixin, UserRealNameFilter
from .models import EmotionRecord
from rangefilter.filters import DateRangeFilter


class CreatedAtTimeFilter(admin.SimpleListFilter):
    """创建时间过滤器 - 支持1天、3天、7天、30天快速筛选"""
    
    title = '创建时间'
    parameter_name = 'created_at_time'
    
    def lookups(self, request, model_admin):
        return (
            ('1', '最近1天'),
            ('3', '最近3天'),
            ('7', '最近7天'),
            ('30', '最近30天'),
            ('all', '全部数据'),
        )
    
    def queryset(self, request, queryset):
        if self.value() is None:
            # 默认返回全部数据
            return queryset
        
        if self.value() == 'all':
            return queryset
        
        try:
            days = int(self.value())
            now = timezone.now()
            start_time = now - timedelta(days=days)
            return queryset.filter(created_at__gte=start_time)
        except (ValueError, TypeError):
            return queryset
    
    def choices(self, changelist):
        """重写choices方法，确保默认选中全部数据"""
        yield {
            'selected': self.value() is None or self.value() == 'all',
            'query_string': changelist.get_query_string({self.parameter_name: 'all'}),
            'display': '全部数据',
        }
        for lookup, title in self.lookup_choices:
            if lookup != 'all':  # 跳过all选项，因为我们已经在上面处理了
                yield {
                    'selected': self.value() == str(lookup),
                    'query_string': changelist.get_query_string({self.parameter_name: lookup}),
                    'display': title,
                }


class UserTrackingStatusFilter(admin.SimpleListFilter):
    """用户追踪状态过滤器"""
    
    title = '用户追踪状态'
    parameter_name = 'user_tracking_status'
    
    def lookups(self, request, model_admin):
        return (
            ('tracked', '已追踪'),
            ('untracked', '未追踪'),
            ('all', '全部用户'),
        )
    
    def queryset(self, request, queryset):
        if self.value() is None or self.value() == 'all':
            return queryset
        
        if self.value() == 'tracked':
            tracked_user_ids = list(User.objects.filter(is_tracked=True).values_list('id', flat=True))
            return queryset.filter(user_id__in=tracked_user_ids)
        
        if self.value() == 'untracked':
            untracked_user_ids = list(User.objects.filter(is_tracked=False).values_list('id', flat=True))
            return queryset.filter(user_id__in=untracked_user_ids)
        
        return queryset
    
    def choices(self, changelist):
        """重写choices方法，确保默认选中全部用户"""
        yield {
            'selected': self.value() is None or self.value() == 'all',
            'query_string': changelist.get_query_string({self.parameter_name: 'all'}),
            'display': '全部用户',
        }
        for lookup, title in self.lookup_choices:
            if lookup != 'all':  # 跳过all选项，因为我们已经在上面处理了
                yield {
                    'selected': self.value() == str(lookup),
                    'query_string': changelist.get_query_string({self.parameter_name: lookup}),
                    'display': title,
                }


@admin.register(EmotionRecord)
class EmotionRecordAdmin(UUIDUserAdminMixin, admin.ModelAdmin):
    """情绪记录管理"""
    
    list_display = [
        'id', 'user_real_name', 'record_date', 'period',
        'depression', 'anxiety', 'energy', 'sleep',
        'mainMood', 'moodIntensity', 'moodSupplementTags_display',
        'moodSupplementText_short', 'created_at'
    ]
    
    list_filter = [
        'period',
        ('record_date', DateRangeFilter),
        CreatedAtTimeFilter,
        UserRealNameFilter,
        UserTrackingStatusFilter,
        'mainMood'
    ]
    
    search_fields = [
        'mainMood', 'mainMoodOther', 'moodSupplementText'
    ]
    
    readonly_fields = [
        'id', 'user_id', 'created_at', 'started_at'
    ]
    
    fieldsets = (
        ('基本信息', {
            'fields': ('user_id', 'record_date', 'period')
        }),
        ('情绪评分', {
            'fields': ('depression', 'anxiety', 'energy', 'sleep')
        }),
        ('情绪描述', {
            'fields': ('mainMood', 'moodIntensity', 'mainMoodOther')
        }),
        ('补充信息', {
            'fields': ('moodSupplementTags', 'moodSupplementText')
        }),
        ('时间信息', {
            'fields': ('started_at', 'created_at'),
            'classes': ('collapse',)
        })
    )
    
    def get_queryset(self, request):
        """优化查询"""
        return super().get_queryset(request).select_related()
    
    def mood_summary(self, obj):
        """情绪摘要"""
        mood_info = []
        if obj.mainMood:
            mood_info.append(f"主要情绪: {obj.mainMood}")
        if obj.moodIntensity:
            mood_info.append(f"强度: {obj.moodIntensity}")
        if obj.moodSupplementText:
            mood_info.append(f"补充: {obj.moodSupplementText[:50]}...")
        
        return " | ".join(mood_info) if mood_info else "无描述"
    
    mood_summary.short_description = '情绪摘要'

    def moodSupplementTags_display(self, obj):
        """情绪补充标签显示"""
        if obj.moodSupplementTags and isinstance(obj.moodSupplementTags, list):
            return ', '.join(obj.moodSupplementTags)
        return '-'
    
    moodSupplementTags_display.short_description = '情绪标签'

    def moodSupplementText_short(self, obj):
        """情绪补充说明简短显示"""
        if obj.moodSupplementText:
            return obj.moodSupplementText[:30] + '...' if len(obj.moodSupplementText) > 30 else obj.moodSupplementText
        return '-'
    
    moodSupplementText_short.short_description = '补充说明'
    
    def score_summary(self, obj):
        """评分摘要"""
        scores = []
        if obj.depression is not None:
            scores.append(f"抑郁: {obj.depression}")
        if obj.anxiety is not None:
            scores.append(f"焦虑: {obj.anxiety}")
        if obj.energy is not None:
            scores.append(f"精力: {obj.energy}")
        if obj.sleep is not None:
            scores.append(f"睡眠: {obj.sleep}")
        
        return " | ".join(scores) if scores else "无评分"
    
    score_summary.short_description = '评分摘要'
    
    def has_add_permission(self, request):
        """禁止手动添加情绪记录"""
        return False
    
    actions = ['export_emotion_records']
    
    def export_emotion_records(self, request, queryset):
        """导出当前筛选的情绪记录（XLSX格式）"""
        try:
            # 使用当前筛选的queryset，而不是忽略它
            emotion_records = queryset.order_by('user_id', 'record_date', 'period')
            
            if not emotion_records.exists():
                self.message_user(request, '当前没有符合条件的情绪记录', level=messages.WARNING)
                return
            
            # 创建 XLSX 文件
            output = io.BytesIO()
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = '情绪记录'
            
            # 定义样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            header_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 设置列宽
            column_widths = [8, 15, 15, 8, 15, 15, 15, 15, 10, 12, 10, 10, 10, 10, 10, 15, 12, 20, 25]
            for i, width in enumerate(column_widths, 1):
                worksheet.column_dimensions[get_column_letter(i)].width = width
            
            # 写入表头
            headers = [
                '编号', '姓名', '调查日期', '时段', '日期-时段',
                '计划时间', '开始时间', '记录时间', '响应延迟(分钟)', '是否按时(1是0否)',
                '答题用时(分钟)', '抑郁水平（1-4）', '焦虑水平（0-10）', '精力得分（0-3）',
                '睡眠得分（0-4）', '主要情绪', '情绪强度（1轻微2中等3明显）',
                '情绪标签', '补充说明'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border
            
            # 写入数据
            from django.utils.timezone import localtime
            for row, record in enumerate(emotion_records, 1):
                # 获取用户信息
                try:
                    user = User.objects.get(id=record.user_id)
                    user_name = user.real_name or user.username
                    # 添加分组信息（如果有）
                    if user.group:
                        user_name = f"{user_name}（{user.group}）"
                except User.DoesNotExist:
                    user_name = "未知用户"
                
                # 格式化日期
                survey_date = record.record_date.strftime('%Y年%m月%d号') if record.record_date else ''
                
                # 时段
                period_map = {'morning': '早', 'evening': '晚'}
                period = period_map.get(record.period, record.period) if record.period else ''
                
                # 日期-时段组合
                date_period = f"{survey_date} {period}" if survey_date and period else ''
                
                # 计划时间（根据时段推断，包含日期）
                planned_datetime = None
                planned_time_str = ''
                if record.record_date and record.period:
                    from datetime import time as datetime_time
                    planned_hour = 8 if record.period == 'morning' else 20
                    # 创建带时区信息的datetime对象
                    naive_datetime = datetime.combine(
                        record.record_date,
                        datetime_time(planned_hour, 0)
                    )
                    planned_datetime = timezone.make_aware(naive_datetime)
                    planned_time_str = localtime(planned_datetime).strftime('%Y年%m月%d日 %H:%M')
                
                # 开始时间
                start_time = localtime(record.started_at).strftime('%Y年%m月%d日 %H:%M') if record.started_at else ''
                
                # 记录时间
                record_time = localtime(record.created_at).strftime('%Y年%m月%d日 %H:%M') if record.created_at else ''
                
                # 响应延迟（分钟）- 计划时间与记录时间的差值
                # 负值表示提早作答，正值表示延迟作答
                delay_minutes = 0  # 默认值改为0而不是空字符串
                
                if record.created_at and planned_datetime:
                    try:
                        # 计算延迟（记录时间 - 计划时间）
                        # 确保两个datetime都有时区信息
                        delay_seconds = (record.created_at - planned_datetime).total_seconds()
                        delay_minutes = int(delay_seconds / 60)
                        
                    except Exception as e:
                        # 如果计算出错，使用默认值
                        delay_minutes = 0
                        print(f"计算延迟时出错: {e}")
                
                # 是否按时（在计划时间前后60分钟内作答算按时）
                on_time = 1 if -60 <= delay_minutes <= 60 else 0
                
                # 答题用时(分钟) - 开始作答到完成记录的时间
                answer_duration = 0
                if record.started_at and record.created_at:
                    answer_seconds = (record.created_at - record.started_at).total_seconds()
                    answer_duration = int(answer_seconds / 60)
                
                # 抑郁水平（1-4分制，根据原始分数转换）
                depression_level = min(4, max(1, int(record.depression / 10) + 1)) if record.depression is not None else ''
                
                # 焦虑水平（保持原始0-10分）
                anxiety_level = record.anxiety if record.anxiety is not None else ''
                
                # 精力得分（保持原始0-3分）
                energy_score = record.energy if record.energy is not None else ''
                
                # 睡眠得分（保持原始0-4分）
                sleep_score = record.sleep if record.sleep is not None else ''
                
                # 主要情绪
                main_mood = record.mainMood if record.mainMood else ''
                
                # 情绪强度（1-3分制）
                mood_intensity_map = {1: '1', 2: '2', 3: '3'}
                mood_intensity = mood_intensity_map.get(record.moodIntensity, '') if record.moodIntensity else ''
                
                # 情绪标签（JSON字段转换为字符串列表）
                mood_tags = ''
                if record.moodSupplementTags and isinstance(record.moodSupplementTags, list):
                    mood_tags = ','.join(record.moodSupplementTags)
                
                # 补充说明
                supplement_text = record.moodSupplementText if record.moodSupplementText else ''
                
                # 写入数据行
                data_row = [
                    row,  # 编号
                    user_name,  # 姓名
                    survey_date,  # 调查日期
                    period,  # 时段
                    date_period,  # 日期-时段组合
                    planned_time_str,  # 计划时间
                    start_time,  # 开始时间
                    record_time,  # 记录时间
                    delay_minutes,  # 响应延迟
                    on_time,  # 是否按时
                    answer_duration,  # 答题用时
                    depression_level,  # 抑郁水平
                    anxiety_level,  # 焦虑水平
                    energy_score,  # 精力得分
                    sleep_score,  # 睡眠得分
                    main_mood,  # 主要情绪
                    mood_intensity,  # 情绪强度
                    mood_tags,  # 情绪标签
                    supplement_text  # 补充说明
                ]
                
                for col, value in enumerate(data_row, 1):
                    cell = worksheet.cell(row=row + 1, column=col, value=value)
                    cell.alignment = cell_alignment
                    cell.border = cell_border
                    # 数字列设置特殊格式
                    if col - 1 in [8, 9, 10, 11, 12, 13, 14, 15, 17]:  # 数字列 (0索引)
                        cell.number_format = '0'
            
            # 设置行高
            for row in range(2, worksheet.max_row + 1):
                worksheet.row_dimensions[row].height = 25
            
            # 冻结首行
            worksheet.freeze_panes = 'A2'
            
            # 保存工作簿到内存
            workbook.save(output)
            output.seek(0)
            
            # 准备响应
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            filename = f'情绪记录_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            self.message_user(request, f'成功导出 {emotion_records.count()} 条情绪记录')
            return response
            
        except Exception as e:
            self.message_user(request, f'导出失败：{str(e)}', level=messages.ERROR)
            return
    
    export_emotion_records.short_description = '导出情绪记录(XLSX)'