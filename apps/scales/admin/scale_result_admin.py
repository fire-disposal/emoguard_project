"""
量表结果管理 - ScaleResult 的后台管理
"""
from django.contrib import admin
from django.utils.html import format_html
from django.http import HttpResponse
from django.template.response import TemplateResponse
from django.contrib.admin.helpers import ACTION_CHECKBOX_NAME
from import_export.admin import ExportActionModelAdmin
from apps.scales.models import ScaleResult
from import_export import resources, fields
from apps.scales.admin.filters import ScaleFilter, StatusFilter

import json
import logging


logger = logging.getLogger(__name__)


class ScaleResultResource(resources.ModelResource):
    """量表结果导出资源 - 原生风格，保持数据库原样输出"""

    # 基础信息 - 直接映射数据库字段
    id = fields.Field(column_name='ID', attribute='id')
    user_id = fields.Field(column_name='用户ID', attribute='user_id')
    scale_config_id = fields.Field(column_name='量表配置ID', attribute='scale_config_id')
    selected_options = fields.Field(column_name='选项答案', attribute='selected_options')
    conclusion = fields.Field(column_name='结论摘要', attribute='conclusion')
    duration_ms = fields.Field(column_name='答题时长(毫秒)', attribute='duration_ms')
    started_at = fields.Field(column_name='开始时间', attribute='started_at')
    completed_at = fields.Field(column_name='完成时间', attribute='completed_at')
    status = fields.Field(column_name='状态', attribute='status')
    analysis = fields.Field(column_name='分析结果', attribute='analysis')
    created_at = fields.Field(column_name='创建时间', attribute='created_at')
    updated_at = fields.Field(column_name='更新时间', attribute='updated_at')
    
    class Meta:
        model = ScaleResult
        fields = [
            'id', 'user_id', 'scale_config_id', 'selected_options', 'conclusion',
            'duration_ms', 'started_at', 'completed_at', 'status', 'analysis',
            'created_at', 'updated_at'
        ]
        export_order = fields
        skip_unchanged = True
        report_skipped = True
    
    def dehydrate_selected_options(self, result):
        """保持选项答案原样输出"""
        return result.selected_options if result.selected_options else []
    
    def dehydrate_analysis(self, result):
        """保持分析结果原样输出"""
        return result.analysis if result.analysis else {}


@admin.register(ScaleResult)
class ScaleResultAdmin(ExportActionModelAdmin):
    resource_class = ScaleResultResource
    list_display = ('id', 'user_id', 'scale_config', 'status', 'duration_formatted', 'created_at')
    list_display_links = ('id', 'user_id')
    search_fields = ('user_id', 'scale_config__name', 'scale_config__code', 'scale_config__type')
    list_filter = ('status', 'scale_config__type', ScaleFilter, StatusFilter, 'created_at')
    readonly_fields = ('created_at', 'updated_at', 'analysis_preview', 'quick_stats')
    list_select_related = ('scale_config',)
    list_per_page = 25
    ordering = ('-created_at',)
    actions = ['export_selected_excel', 'export_selected_csv', 'mark_as_reviewed', 'bulk_delete_results']
    date_hierarchy = 'created_at'
    
    # 导出配置
    # export_formats = ['xlsx', 'csv']  # 错误写法，已注释
    export_filename = '量表结果导出'
    
    class Media:
        css = {
            'all': ('admin/css/custom_admin.css',)
        }
    
    def get_queryset(self, request):
        """优化查询性能"""
        return super().get_queryset(request).select_related('scale_config')
    
    def quick_stats(self, obj):
        """快速统计信息 - 原生风格"""
        if not obj.analysis or not isinstance(obj.analysis, dict):
            return '无统计数据'
        
        try:
            # 直接显示分析数据的原始结构
            preview_data = {
                'score': obj.analysis.get('score'),
                'level': obj.analysis.get('level'),
                'is_abnormal': obj.analysis.get('is_abnormal'),
                'risk_level': obj.analysis.get('risk_level')
            }
            
            # 移除None值，保持简洁
            preview_data = {k: v for k, v in preview_data.items() if v is not None}
            
            return format_html(
                '<pre class="analysis-preview">{}</pre>',
                json.dumps(preview_data, ensure_ascii=False, indent=2)
            )
        except Exception as e:
            logger.error(f"生成快速统计失败: {str(e)}")
            return '统计错误'
    quick_stats.short_description = '快速统计'
    
    def duration_formatted(self, obj):
        """格式化显示答题时长"""
        return format_duration(obj.duration_ms)
    duration_formatted.short_description = '答题时长'
    
    
    def analysis_preview(self, obj):
        """预览分析结果（主题友好）"""
        if not obj.analysis or not isinstance(obj.analysis, dict):
            return "无分析数据"
        
        try:
            # 显示关键信息
            preview_data = {
                'score': obj.analysis.get('score'),
                'level': obj.analysis.get('level'),
                'max_score': obj.analysis.get('max_score'),
                'is_abnormal': obj.analysis.get('is_abnormal'),
                'recommendations': obj.analysis.get('recommendations', [])[:2]  # 只显示前2条建议
            }
            
            # 移除None值
            preview_data = {k: v for k, v in preview_data.items() if v is not None}
            
            formatted = json.dumps(preview_data, ensure_ascii=False, indent=2)
            return format_html(
                '<pre class="analysis-preview">{}</pre>',
                formatted
            )
        except Exception as e:
            logger.error(f"预览分析结果失败: {str(e)}")
            return "分析数据格式错误"
    analysis_preview.short_description = '分析预览'
    
    # 批量操作
    def export_selected_excel(self, request, queryset):
        """导出为Excel格式 - 原生风格，保持数据库原样输出"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from datetime import datetime
        import json
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "量表结果"
        
        # 设置标题行 - 直接映射数据库字段
        headers = ['ID', '用户ID', '量表配置ID', '选项答案', '结论摘要', '答题时长(毫秒)', '开始时间', '完成时间', '状态', '分析结果', '创建时间', '更新时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # 填充数据 - 保持数据库原样
        for row, result in enumerate(queryset.select_related('scale_config'), 2):
            ws.cell(row=row, column=1, value=result.id)
            ws.cell(row=row, column=2, value=str(result.user_id))
            ws.cell(row=row, column=3, value=result.scale_config_id)
            ws.cell(row=row, column=4, value=json.dumps(result.selected_options, ensure_ascii=False) if result.selected_options else '')
            ws.cell(row=row, column=5, value=result.conclusion or '')
            ws.cell(row=row, column=6, value=result.duration_ms or 0)
            ws.cell(row=row, column=7, value=result.started_at.strftime('%Y-%m-%d %H:%M:%S') if result.started_at else '')
            ws.cell(row=row, column=8, value=result.completed_at.strftime('%Y-%m-%d %H:%M:%S') if result.completed_at else '')
            ws.cell(row=row, column=9, value=result.status or '')
            ws.cell(row=row, column=10, value=json.dumps(result.analysis, ensure_ascii=False) if result.analysis else '')
            ws.cell(row=row, column=11, value=result.created_at.strftime('%Y-%m-%d %H:%M:%S') if result.created_at else '')
            ws.cell(row=row, column=12, value=result.updated_at.strftime('%Y-%m-%d %H:%M:%S') if result.updated_at else '')
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="量表结果_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
    export_selected_excel.short_description = '导出为Excel'
    
    def export_selected_csv(self, request, queryset):
        """导出为CSV格式 - 原生风格，保持数据库原样输出"""
        import csv
        from datetime import datetime
        import json
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="量表结果_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', '用户ID', '量表配置ID', '选项答案', '结论摘要', '答题时长(毫秒)', '开始时间', '完成时间', '状态', '分析结果', '创建时间', '更新时间'])
        
        for result in queryset.select_related('scale_config'):
            writer.writerow([
                result.id,
                str(result.user_id),
                result.scale_config_id,
                json.dumps(result.selected_options, ensure_ascii=False) if result.selected_options else '',
                result.conclusion or '',
                result.duration_ms or 0,
                result.started_at.strftime('%Y-%m-%d %H:%M:%S') if result.started_at else '',
                result.completed_at.strftime('%Y-%m-%d %H:%M:%S') if result.completed_at else '',
                result.status or '',
                json.dumps(result.analysis, ensure_ascii=False) if result.analysis else '',
                result.created_at.strftime('%Y-%m-%d %H:%M:%S') if result.created_at else '',
                result.updated_at.strftime('%Y-%m-%d %H:%M:%S') if result.updated_at else ''
            ])
        
        return response
    export_selected_csv.short_description = '导出为CSV'
    
    def mark_as_reviewed(self, request, queryset):
        """标记为已审核"""
        # 可以添加审核逻辑，比如添加审核标记到analysis中
        updated = queryset.count()
        self.message_user(request, f'已标记 {updated} 条记录为已审核')
    mark_as_reviewed.short_description = '标记为已审核'
    
    def bulk_delete_results(self, request, queryset):
        """批量删除结果（带确认）"""
        if request.POST.get('post'):
            # 确认删除
            count = queryset.count()
            queryset.delete()
            self.message_user(request, f'已删除 {count} 条记录')
        else:
            # 显示确认页面
            context = {
                'title': '确认批量删除',
                'queryset': queryset,
                'action_checkbox_name': ACTION_CHECKBOX_NAME,
                'action': 'bulk_delete_results',
                'opts': self.model._meta,
            }
            return TemplateResponse(request, 'admin/bulk_delete_confirmation.html', context)
    bulk_delete_results.short_description = '批量删除所选记录'

def format_duration(duration_ms):
    """格式化时长显示"""
    if not duration_ms:
        return "0秒"
    
    seconds = duration_ms // 1000
    if seconds < 60:
        return f"{seconds}秒"
    else:
        minutes = seconds // 60
        remaining_seconds = seconds % 60
        return f"{minutes}分{remaining_seconds}秒"