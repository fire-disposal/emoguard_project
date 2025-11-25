"""
人口学信息导出工具模块
"""
from apps.users.models import User

def get_demographic_info(user_id):
    try:
        user = User.objects.get(id=user_id)
        return {
            "real_name": user.real_name or "未知",
            "gender": {
                "male": "男", "female": "女", "other": "其他",
                "男": "男", "女": "女", "其他": "其他"
            }.get(user.gender, "未知"),
            "age": user.age or "未知",
            "education": user.education or "未知",
            "province": user.province or "未知",
            "city": user.city or "未知",
            "district": user.district or "未知",
            "phone": user.phone or "未知",
            "score_scd": user.score_scd if hasattr(user, "score_scd") else None,
            "score_mmse": user.score_mmse if hasattr(user, "score_mmse") else None,
            "score_moca": user.score_moca if hasattr(user, "score_moca") else None,
            "score_gad7": user.score_gad7 if hasattr(user, "score_gad7") else None,
            "score_phq9": user.score_phq9 if hasattr(user, "score_phq9") else None,
            "score_adl": user.score_adl if hasattr(user, "score_adl") else None,
        }
    except User.DoesNotExist:
        return {
            "real_name": "用户不存在",
            "gender": "未知",
            "age": "未知",
            "education": "未知",
            "province": "未知",
            "city": "未知",
            "district": "未知",
            "phone": "未知",
            "score_scd": None,
            "score_mmse": None,
            "score_moca": None,
            "score_gad7": None,
            "score_phq9": None,
            "score_adl": None,
        }
def fill_demographic_cells(ws, row, user_info, start_col=3):
    """
    批量填充人口学相关字段到Excel行
    :param ws: 工作表对象
    :param row: 行号
    :param user_info: 人口学信息字典
    :param start_col: 起始列号（默认第3列）
    """
    fields = [
        "real_name", "gender", "age", "education", "province", "city", "district", "phone",
        "score_scd", "score_mmse", "score_moca", "score_gad7", "score_phq9", "score_adl"
    ]
    for idx, key in enumerate(fields):
        ws.cell(row=row, column=start_col + idx, value=user_info.get(key, ""))
def get_demographic_csv_row(user_info):
    """
    构建人口学信息的csv行（列表）
    :param user_info: 人口学信息字典
    :return: 列表
    """
    fields = [
        "real_name", "gender", "age", "education", "province", "city", "district", "phone",
        "score_scd", "score_mmse", "score_moca", "score_gad7", "score_phq9", "score_adl"
    ]
    return [user_info.get(key, "") for key in fields]

def build_row_with_demographics(user_info, extra_fields: dict, extra_field_order: list):
    """
    构建一行数据，人口学信息在前，业务字段按指定顺序追加
    :param user_info: 人口学信息字典
    :param extra_fields: 业务字段字典
    :param extra_field_order: 业务字段顺序列表
    :return: 列表
    """
    demo_row = get_demographic_csv_row(user_info)
    biz_row = [extra_fields.get(k, "") for k in extra_field_order]
    return demo_row + biz_row

# -- CSV导出功能 --

def build_excel_with_demographics(queryset, get_user_id, extra_field_order, extra_field_titles):
    """
    构建带人口学信息的Excel并返回HttpResponse
    :param queryset: 业务数据集
    :param get_user_id: 获取user_id的方法
    :param extra_field_order: 业务字段顺序列表
    :param extra_field_titles: 业务字段标题列表
    :return: HttpResponse
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from django.http import HttpResponse
    from datetime import datetime

    demo_titles = [
        "用户姓名", "性别", "年龄", "学历", "所在省份", "所在城市", "所在区县", "手机号",
        "SCD分数", "MMSE分数", "MoCA分数", "GAD-7分数", "PHQ-9分数", "ADL分数"
    ]
    headers = demo_titles + extra_field_titles
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "导出数据"

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)

    for row, record in enumerate(queryset, 2):
        user_info = get_demographic_info(get_user_id(record))
        extra_fields = {k: getattr(record, k, "") for k in extra_field_order}
        extra_fields['record'] = record
        data_row = build_row_with_demographics(user_info, extra_fields, extra_field_order)
        for col, value in enumerate(data_row, 1):
            import uuid
            if isinstance(value, uuid.UUID):
                value = str(value)
            if isinstance(value, datetime) and getattr(value, "tzinfo", None) is not None:
                value = value.replace(tzinfo=None)
            ws.cell(row=row, column=col, value=value)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="导出_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )
    wb.save(response)
    return response

# -- CSV导出功能 --

def build_csv_with_demographics(queryset, get_user_id, extra_field_order, extra_field_titles):
    """
    构建带人口学信息的CSV并返回HttpResponse
    :param queryset: 业务数据集
    :param get_user_id: 获取user_id的方法
    :param extra_field_order: 业务字段顺序列表
    :param extra_field_titles: 业务字段标题列表
    :return: HttpResponse
    """
    import csv
    from django.http import HttpResponse
    from datetime import datetime

    demo_titles = [
        "用户姓名", "性别", "年龄", "学历", "所在省份", "所在城市", "所在区县", "手机号",
        "SCD分数", "MMSE分数", "MoCA分数", "GAD-7分数", "PHQ-9分数", "ADL分数"
    ]
    headers = demo_titles + extra_field_titles

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        f'attachment; filename="导出_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    )
    writer = csv.writer(response)
    writer.writerow(headers)

    for record in queryset:
        user_info = get_demographic_info(get_user_id(record))
        extra_fields = {k: getattr(record, k, "") for k in extra_field_order}
        import uuid
        row = build_row_with_demographics(user_info, extra_fields, extra_field_order)
        row = [str(v) if isinstance(v, uuid.UUID) else v for v in row]
        writer.writerow(row)

    return response
