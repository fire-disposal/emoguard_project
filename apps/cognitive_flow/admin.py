"""
认知测评流程管理后台 - 添加数据导出功能
"""
from django.contrib import admin
from django.http import HttpResponse
from django.utils.html import format_html
from import_export.admin import ExportActionModelAdmin
from import_export import resources, fields
from apps.cognitive_flow.models import CognitiveAssessmentRecord
import json
import csv
from datetime import datetime


class CognitiveAssessmentRecordResource(resources.ModelResource):
    """认知测评记录导出资源 - 面向数据分析"""
    
    id = fields.Field(column_name='记录ID', attribute='id')
    user_id_display = fields.Field(column_name='用户ID', attribute='user_id')
    real_name = fields.Field(column_name='用户姓名', attribute='real_name')
    gender = fields.Field(column_name='性别', attribute='gender')
    age = fields.Field(column_name='年龄', attribute='age')
    education = fields.Field(column_name='学历', attribute='education')
    province = fields.Field(column_name='省份', attribute='province')
    city = fields.Field(column_name='城市', attribute='city')
    district = fields.Field(column_name='区县', attribute='district')
    
    # 核心测评得分 - 数据分析重点
    score_scd = fields.Field(column_name='SCD得分(主观认知下降)', attribute='score_scd')
    score_mmse = fields.Field(column_name='MMSE得分(简易智能状态)', attribute='score_mmse')
    score_moca = fields.Field(column_name='MoCA得分(蒙特利尔认知)', attribute='score_moca')
    score_gad7 = fields.Field(column_name='GAD7得分(焦虑筛查)', attribute='score_gad7')
    score_phq9 = fields.Field(column_name='PHQ9得分(抑郁筛查)', attribute='score_phq9')
    score_adl = fields.Field(column_name='ADL得分(日常生活能力)', attribute='score_adl')
    
    # 分析结果
    cognitive_status = fields.Field(column_name='认知状态评估', attribute='analysis')
    anxiety_level = fields.Field(column_name='焦虑水平', attribute='analysis')
    depression_level = fields.Field(column_name='抑郁水平', attribute='analysis')
    daily_function = fields.Field(column_name='日常功能评估', attribute='analysis')
    
    # 时间信息
    started_at = fields.Field(column_name='开始时间', attribute='started_at')
    completed_at = fields.Field(column_name='完成时间', attribute='completed_at')
    duration_minutes = fields.Field(column_name='测评时长(分钟)', attribute='duration_ms')
    created_at = fields.Field(column_name='记录时间', attribute='created_at')
    
    class Meta:
        model = CognitiveAssessmentRecord
        fields = [
            'id', 'user_id_display', 'real_name', 'gender', 'age', 'education',
            'province', 'city', 'district',
            'score_scd', 'score_mmse', 'score_moca', 'score_gad7', 'score_phq9', 'score_adl',
            'cognitive_status', 'anxiety_level', 'depression_level', 'daily_function',
            'duration_minutes', 'started_at', 'completed_at', 'created_at'
        ]
        export_order = fields
        skip_unchanged = True
    
    def dehydrate_user_id_display(self, record):
        """用户ID显示（截断）"""
        return str(record.user_id)[:8]
    
    def dehydrate_gender(self, record):
        """性别显示优化"""
        gender_map = {
            'male': '男',
            'female': '女',
            'other': '其他',
            '': '未知'
        }
        return gender_map.get(record.gender, record.gender or '未知')
    
    def dehydrate_age(self, record):
        """年龄显示"""
        return record.age if record.age is not None else '未知'
    
    def dehydrate_education(self, record):
        """学历显示"""
        education_map = {
            'primary': '小学',
            'junior_high': '初中',
            'senior_high': '高中',
            'college': '大专',
            'bachelor': '本科',
            'master': '硕士',
            'doctor': '博士',
            '': '未知'
        }
        return education_map.get(record.education, record.education or '未知')
    
    def dehydrate_cognitive_status(self, record):
        """认知状态评估"""
        if not record.analysis or not isinstance(record.analysis, dict):
            return '未评估'
        
        # 基于MMSE和MoCA得分评估认知状态
        mmse_score = record.score_mmse
        moca_score = record.score_moca
        
        if mmse_score is not None and moca_score is not None:
            if mmse_score >= 27 and moca_score >= 26:
                return '认知正常'
            elif mmse_score >= 24 and moca_score >= 18:
                return '轻度认知障碍'
            else:
                return '认知障碍'
        elif mmse_score is not None:
            if mmse_score >= 27:
                return '认知正常'
            elif mmse_score >= 24:
                return '轻度认知障碍'
            else:
                return '认知障碍'
        elif moca_score is not None:
            if moca_score >= 26:
                return '认知正常'
            elif moca_score >= 18:
                return '轻度认知障碍'
            else:
                return '认知障碍'
        
        return '数据不足'
    
    def dehydrate_anxiety_level(self, record):
        """焦虑水平评估"""
        gad7_score = record.score_gad7
        if gad7_score is None:
            return '未评估'
        
        if gad7_score <= 4:
            return '无焦虑'
        elif gad7_score <= 9:
            return '轻度焦虑'
        elif gad7_score <= 14:
            return '中度焦虑'
        else:
            return '重度焦虑'
    
    def dehydrate_depression_level(self, record):
        """抑郁水平评估"""
        phq9_score = record.score_phq9
        if phq9_score is None:
            return '未评估'
        
        if phq9_score <= 4:
            return '无抑郁'
        elif phq9_score <= 9:
            return '轻度抑郁'
        elif phq9_score <= 14:
            return '中度抑郁'
        elif phq9_score <= 19:
            return '中重度抑郁'
        else:
            return '重度抑郁'
    
    def dehydrate_daily_function(self, record):
        """日常功能评估"""
        adl_score = record.score_adl
        if adl_score is None:
            return '未评估'
        
        # 假设ADL得分范围，根据实际情况调整
        if adl_score <= 16:
            return '功能正常'
        elif adl_score <= 20:
            return '轻度功能障碍'
        elif adl_score <= 24:
            return '中度功能障碍'
        else:
            return '重度功能障碍'
    
    def dehydrate_duration_minutes(self, record):
        """测评时长转换为分钟"""
        if record.duration_ms:
            duration_seconds = (record.completed_at - record.started_at).total_seconds()
            return round(duration_seconds / 60, 1)
        return '未知'


@admin.register(CognitiveAssessmentRecord)
class CognitiveAssessmentRecordAdmin(ExportActionModelAdmin):
    resource_class = CognitiveAssessmentRecordResource
    list_display = (
        "id", "user_info", "basic_info", "cognitive_scores", 
        "psychological_status", "duration_info", "created_at"
    )
    list_display_links = ("id", "user_info")
    search_fields = ("user_id", "real_name", "phone")
    list_filter = ("province", "city", "education", "gender", "created_at")
    readonly_fields = ("started_at", "completed_at", "analysis_preview")
    ordering = ("-started_at",)
    actions = ['export_selected_excel', 'export_selected_csv', 'export_analysis_summary']
    date_hierarchy = 'created_at'
    
    def user_info(self, obj):
        """显示用户信息"""
        return format_html(
            '<div style="display: flex; flex-direction: column;">'
            '<span style="font-weight: 600;">{}</span>'
            '<span style="font-size: 12px; color: #999;">ID: {}</span>'
            '<span style="font-size: 12px; color: #999;">{}</span>'
            '</div>',
            obj.real_name or '未知用户',
            str(obj.user_id)[:8] + '...',
            obj.phone or '无电话'
        )
    user_info.short_description = '用户信息'
    
    def basic_info(self, obj):
        """显示基本信息"""
        gender_display = {'male': '男', 'female': '女', 'other': '其他'}.get(obj.gender, obj.gender or '未知')
        
        return format_html(
            '<div style="display: flex; flex-direction: column;">'
            '<span>{}岁 | {}</span>'
            '<span>{}</span>'
            '<span>{} {}</span>'
            '</div>',
            obj.age if obj.age else '未知',
            gender_display,
            obj.education or '未知学历',
            obj.province or '',
            obj.city or ''
        )
    basic_info.short_description = '基本信息'
    
    def cognitive_scores(self, obj):
        """显示认知测评得分"""
        scores = []
        if obj.score_mmse is not None:
            scores.append(f'MMSE: {obj.score_mmse}')
        if obj.score_moca is not None:
            scores.append(f'MoCA: {obj.score_moca}')
        if obj.score_scd is not None:
            scores.append(f'SCD: {obj.score_scd}')
        
        return format_html(
            '<div style="display: flex; flex-direction: column;">'
            '{}'
            '</div>',
            '<br>'.join(scores) if scores else '无认知数据'
        )
    cognitive_scores.short_description = '认知得分'
    
    def psychological_status(self, obj):
        """显示心理状态"""
        status_parts = []
        
        # 焦虑状态
        if obj.score_gad7 is not None:
            if obj.score_gad7 <= 4:
                anxiety = '无焦虑'
            elif obj.score_gad7 <= 9:
                anxiety = '轻度焦虑'
            elif obj.score_gad7 <= 14:
                anxiety = '中度焦虑'
            else:
                anxiety = '重度焦虑'
            status_parts.append(f'GAD7: {anxiety}')
        
        # 抑郁状态
        if obj.score_phq9 is not None:
            if obj.score_phq9 <= 4:
                depression = '无抑郁'
            elif obj.score_phq9 <= 9:
                depression = '轻度抑郁'
            elif obj.score_phq9 <= 14:
                depression = '中度抑郁'
            elif obj.score_phq9 <= 19:
                depression = '中重度抑郁'
            else:
                depression = '重度抑郁'
            status_parts.append(f'PHQ9: {depression}')
        
        # 日常功能
        if obj.score_adl is not None:
            if obj.score_adl <= 16:
                daily_func = '功能正常'
            elif obj.score_adl <= 20:
                daily_func = '轻度障碍'
            elif obj.score_adl <= 24:
                daily_func = '中度障碍'
            else:
                daily_func = '重度障碍'
            status_parts.append(f'ADL: {daily_func}')
        
        return format_html(
            '<div style="display: flex; flex-direction: column;">'
            '{}'
            '</div>',
            '<br>'.join(status_parts) if status_parts else '无心理数据'
        )
    psychological_status.short_description = '心理状态'
    
    def duration_info(self, obj):
        """显示时长信息"""
        if obj.started_at and obj.completed_at:
            duration = obj.completed_at - obj.started_at
            minutes = duration.total_seconds() / 60
            return format_html(
                '<div style="display: flex; flex-direction: column;">'
                '<span>{}分钟</span>'
                '<span style="font-size: 12px; color: #999;">{}</span>'
                '</div>',
                round(minutes, 1),
                obj.completed_at.strftime('%m月%d日 %H:%M')
            )
        return '未完成'
    duration_info.short_description = '测评时长'
    
    def analysis_preview(self, obj):
        """分析结果预览"""
        analysis_data = {
            'SCD得分': obj.score_scd,
            'MMSE得分': obj.score_mmse,
            'MoCA得分': obj.score_moca,
            'GAD7得分': obj.score_gad7,
            'PHQ9得分': obj.score_phq9,
            'ADL得分': obj.score_adl,
            '分析数据': obj.analysis if obj.analysis else {}
        }
        
        # 移除None值
        analysis_data = {k: v for k, v in analysis_data.items() if v is not None}
        
        formatted = json.dumps(analysis_data, ensure_ascii=False, indent=2)
        return format_html('<pre class="analysis-preview">{}</pre>', formatted)
    analysis_preview.short_description = '数据预览'
    
    def export_selected_excel(self, request, queryset):
        """导出认知测评记录为Excel格式"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "认知测评数据"
        
        # 设置标题行 - 面向数据分析的字段
        headers = [
            '记录ID', '用户ID', '姓名', '性别', '年龄', '学历',
            '省份', '城市', '区县',
            'SCD得分', 'MMSE得分', 'MoCA得分', 'GAD7得分', 'PHQ9得分', 'ADL得分',
            '认知状态', '焦虑水平', '抑郁水平', '日常功能',
            '测评时长(分钟)', '开始时间', '完成时间', '记录时间'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 填充数据
        for row, record in enumerate(queryset, 2):
            # 计算各项评估
            cognitive_status = self.get_cognitive_status(record)
            anxiety_level = self.get_anxiety_level(record)
            depression_level = self.get_depression_level(record)
            daily_function = self.get_daily_function(record)
            duration_minutes = self.get_duration_minutes(record)
            
            ws.cell(row=row, column=1, value=record.id)
            ws.cell(row=row, column=2, value=str(record.user_id)[:8])
            ws.cell(row=row, column=3, value=record.real_name or '未知')
            ws.cell(row=row, column=4, value=self.get_gender_display(record.gender))
            ws.cell(row=row, column=5, value=record.age if record.age else '未知')
            ws.cell(row=row, column=6, value=self.get_education_display(record.education))
            ws.cell(row=row, column=7, value=record.province or '未知')
            ws.cell(row=row, column=8, value=record.city or '未知')
            ws.cell(row=row, column=9, value=record.district or '未知')
            ws.cell(row=row, column=10, value=record.score_scd if record.score_scd is not None else 'N/A')
            ws.cell(row=row, column=11, value=record.score_mmse if record.score_mmse is not None else 'N/A')
            ws.cell(row=row, column=12, value=record.score_moca if record.score_moca is not None else 'N/A')
            ws.cell(row=row, column=13, value=record.score_gad7 if record.score_gad7 is not None else 'N/A')
            ws.cell(row=row, column=14, value=record.score_phq9 if record.score_phq9 is not None else 'N/A')
            ws.cell(row=row, column=15, value=record.score_adl if record.score_adl is not None else 'N/A')
            ws.cell(row=row, column=16, value=cognitive_status)
            ws.cell(row=row, column=17, value=anxiety_level)
            ws.cell(row=row, column=18, value=depression_level)
            ws.cell(row=row, column=19, value=daily_function)
            ws.cell(row=row, column=20, value=duration_minutes)
            ws.cell(row=row, column=21, value=record.started_at.strftime('%Y-%m-%d %H:%M') if record.started_at else 'N/A')
            ws.cell(row=row, column=22, value=record.completed_at.strftime('%Y-%m-%d %H:%M') if record.completed_at else 'N/A')
            ws.cell(row=row, column=23, value=record.created_at.strftime('%Y-%m-%d %H:%M'))
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="认知测评数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
    export_selected_excel.short_description = '导出Excel(数据分析版)'
    
    def export_selected_csv(self, request, queryset):
        """导出认知测评记录为CSV格式"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="认知测评数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            '记录ID', '用户ID', '姓名', '性别', '年龄', '学历',
            '省份', '城市', '区县',
            'SCD得分', 'MMSE得分', 'MoCA得分', 'GAD7得分', 'PHQ9得分', 'ADL得分',
            '认知状态', '焦虑水平', '抑郁水平', '日常功能',
            '测评时长(分钟)', '开始时间', '完成时间', '记录时间'
        ])
        
        for record in queryset:
            writer.writerow([
                record.id,
                str(record.user_id)[:8],
                record.real_name or '未知',
                self.get_gender_display(record.gender),
                record.age if record.age else '未知',
                self.get_education_display(record.education),
                record.province or '未知',
                record.city or '未知',
                record.district or '未知',
                record.score_scd if record.score_scd is not None else 'N/A',
                record.score_mmse if record.score_mmse is not None else 'N/A',
                record.score_moca if record.score_moca is not None else 'N/A',
                record.score_gad7 if record.score_gad7 is not None else 'N/A',
                record.score_phq9 if record.score_phq9 is not None else 'N/A',
                record.score_adl if record.score_adl is not None else 'N/A',
                self.get_cognitive_status(record),
                self.get_anxiety_level(record),
                self.get_depression_level(record),
                self.get_daily_function(record),
                self.get_duration_minutes(record),
                record.started_at.strftime('%Y-%m-%d %H:%M') if record.started_at else 'N/A',
                record.completed_at.strftime('%Y-%m-%d %H:%M') if record.completed_at else 'N/A',
                record.created_at.strftime('%Y-%m-%d %H:%M')
            ])
        
        return response
    export_selected_csv.short_description = '导出CSV(数据分析版)'
    
    def export_analysis_summary(self, request, queryset):
        """导出分析汇总"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="认知测评分析汇总_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['统计项目', '数值', '说明'])
        
        total_count = queryset.count()
        writer.writerow(['总记录数', total_count, ''])
        
        # 性别分布
        gender_stats = {}
        for record in queryset:
            gender = self.get_gender_display(record.gender)
            gender_stats[gender] = gender_stats.get(gender, 0) + 1
        
        for gender, count in gender_stats.items():
            writer.writerow([f'性别-{gender}', count, f'占比: {count/total_count*100:.1f}%'])
        
        # 年龄分布
        age_groups = {'<30': 0, '30-39': 0, '40-49': 0, '50-59': 0, '60-69': 0, '70+': 0}
        for record in queryset:
            if record.age:
                if record.age < 30:
                    age_groups['<30'] += 1
                elif record.age < 40:
                    age_groups['30-39'] += 1
                elif record.age < 50:
                    age_groups['40-49'] += 1
                elif record.age < 60:
                    age_groups['50-59'] += 1
                elif record.age < 70:
                    age_groups['60-69'] += 1
                else:
                    age_groups['70+'] += 1
        
        for group, count in age_groups.items():
            if count > 0:
                writer.writerow([f'年龄组-{group}', count, f'占比: {count/total_count*100:.1f}%'])
        
        # 认知状态分布
        cognitive_stats = {}
        for record in queryset:
            status = self.get_cognitive_status(record)
            cognitive_stats[status] = cognitive_stats.get(status, 0) + 1
        
        for status, count in cognitive_stats.items():
            writer.writerow([f'认知状态-{status}', count, f'占比: {count/total_count*100:.1f}%'])
        
        return response
    export_analysis_summary.short_description = '导出分析汇总'
    
    # 辅助方法
    def get_gender_display(self, gender):
        """性别显示"""
        gender_map = {
            'male': '男',
            'female': '女',
            'other': '其他',
            '': '未知'
        }
        return gender_map.get(gender, gender or '未知')
    
    def get_education_display(self, education):
        """学历显示"""
        education_map = {
            'primary': '小学',
            'junior_high': '初中',
            'senior_high': '高中',
            'college': '大专',
            'bachelor': '本科',
            'master': '硕士',
            'doctor': '博士',
            '': '未知'
        }
        return education_map.get(education, education or '未知')
    
    def get_cognitive_status(self, record):
        """获取认知状态"""
        mmse_score = record.score_mmse
        moca_score = record.score_moca
        
        if mmse_score is not None and moca_score is not None:
            if mmse_score >= 27 and moca_score >= 26:
                return '认知正常'
            elif mmse_score >= 24 and moca_score >= 18:
                return '轻度认知障碍'
            else:
                return '认知障碍'
        elif mmse_score is not None:
            if mmse_score >= 27:
                return '认知正常'
            elif mmse_score >= 24:
                return '轻度认知障碍'
            else:
                return '认知障碍'
        elif moca_score is not None:
            if moca_score >= 26:
                return '认知正常'
            elif moca_score >= 18:
                return '轻度认知障碍'
            else:
                return '认知障碍'
        
        return '数据不足'
    
    def get_anxiety_level(self, record):
        """获取焦虑水平"""
        gad7_score = record.score_gad7
        if gad7_score is None:
            return '未评估'
        
        if gad7_score <= 4:
            return '无焦虑'
        elif gad7_score <= 9:
            return '轻度焦虑'
        elif gad7_score <= 14:
            return '中度焦虑'
        else:
            return '重度焦虑'
    
    def get_depression_level(self, record):
        """获取抑郁水平"""
        phq9_score = record.score_phq9
        if phq9_score is None:
            return '未评估'
        
        if phq9_score <= 4:
            return '无抑郁'
        elif phq9_score <= 9:
            return '轻度抑郁'
        elif phq9_score <= 14:
            return '中度抑郁'
        elif phq9_score <= 19:
            return '中重度抑郁'
        else:
            return '重度抑郁'
    
    def get_daily_function(self, record):
        """获取日常功能"""
        adl_score = record.score_adl
        if adl_score is None:
            return '未评估'
        
        if adl_score <= 16:
            return '功能正常'
        elif adl_score <= 20:
            return '轻度功能障碍'
        elif adl_score <= 24:
            return '中度功能障碍'
        else:
            return '重度功能障碍'
    
    def get_duration_minutes(self, record):
        """获取测评时长（分钟）"""
        if record.started_at and record.completed_at:
            duration = record.completed_at - record.started_at
            return round(duration.total_seconds() / 60, 1)
        return '未知'